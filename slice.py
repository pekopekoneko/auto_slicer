# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""


special_col = 1  #分类的那个列,就比如 A是1 B是2 C是3 D是4
special_col = special_col-1
from openpyxl import load_workbook
from openpyxl.utils import *
import numpy as np
import pandas as pd
from openpyxl.styles import Font,Side
from openpyxl.styles import Border,colors,Alignment

wb = load_workbook('test.xlsx')
ws = wb[wb.sheetnames[0]]
row_length = len(ws['A'])#获取行长度
col_length = len(ws[1])#获取列长度
fontobj = Font(name=u'宋体', bold=False, italic=False, size=12)#字体设置
border = Border(left=Side(style='thin',color=colors.BLACK),
                right=Side(style='thin',color=colors.BLACK),
                top=Side(style='thin',color=colors.BLACK),
                bottom=Side(style='thin',color=colors.BLACK))#边框设置
this_alignment = Alignment(horizontal='center', vertical='center')#居中设置
content=[]                              #获取列名
for i in range(1,(col_length+1)):
    content.append(get_column_letter(i)) 

def my_get_col(col_num,sheet,start=0):  #获得某一列，存储为list
    result=[]
    col_num = content[col_num]
    temp = sheet[col_num]
    for i in range(start,row_length):
        result.append(temp[i].value)
    return result

def my_get_row(row_num,sheet,start=0): #获得某一行，存储为list
    result=[]
    temp = sheet[row_num]
    for i in range(start,col_length):
        result.append(temp[i].value)
    return result

types = np.unique(my_get_col(special_col,ws,1)) #提取每一种种类标签

col_width=[]
for i in content:                                  #获取列宽
    col_width.append(ws.column_dimensions[i].width)

special_col_index = my_get_col(special_col,ws)   #获取特殊列用于后面的if循环
colname = my_get_row(1,ws)                       #第一行为列名称

for i in types:                              #开始按照种类标签循环
    this_class=[]
    for j in range(1,row_length):           #存储符合标签的内容
        if special_col_index[j] == i:
            this_class.append(my_get_row(j+1,ws))
    df = pd.DataFrame(this_class, columns=colname)#将存好的内容转化为dataframe  
    df.to_excel(str(i)+".xlsx", index=False)#制作好相应的初期excel
    temp_wb = load_workbook(str(i)+".xlsx")#开始对excel进行格式处理
    temp_ws = temp_wb[temp_wb.sheetnames[0]]
    for t in range(1,(col_length+1)):
      temp_ws.column_dimensions[get_column_letter(t)].width = col_width[t-1]
    for irow in range(1,len(temp_ws['A'])+1):#我们只能一个一个单元格来处理
        for icol in content:
            name = str(icol)+str(irow)
            temp_ws[name].font = fontobj#设置字体
            temp_ws[name].border = border#设置边框
            temp_ws[name].alignment = this_alignment
    temp_wb.save(str(i)+".xlsx")#保存输出
        

